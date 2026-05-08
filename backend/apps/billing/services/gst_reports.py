"""GSTR-1 + GSTR-3B report generation.

GSTR-1 — outward supplies (sales). Filed monthly/quarterly. Sections:
  - B2B  : sales to GST-registered customers (rare for hospitals; GSTIN required)
  - B2C  : sales to consumers (most hospital revenue)
           - B2CL: invoice value > ₹2.5 lakh inter-state (per-invoice)
           - B2CS: everything else (state+rate-wise summary)
  - HSN  : HSN code-wise summary (mandatory)
  - DOCS : document issued summary (count of invoices, refunds)

GSTR-3B — monthly summary return. Sections relevant to hospitals:
  - 3.1 (a) Outward taxable supplies (other than zero-rated, nil-rated, exempted)
  - 3.1 (c) Other outward supplies (Nil rated, exempted)
  - 6.1     Tax payable: CGST / SGST / IGST

This module builds JSON suitable for upload to the GSTN portal AND a workbook
for human review. Hospital reports are typically simple — most outpatient and
diagnostic services are GST-exempt; admission/medicines/some procedures attract
GST. The split is driven by InvoiceItem.gst_rate.
"""
from collections import defaultdict
from datetime import date
from decimal import Decimal
from calendar import monthrange

from django.db.models import Sum, F
from apps.billing.models import Invoice, InvoiceItem


def _q(amount):
    return Decimal(amount).quantize(Decimal("0.01"))


def _month_range(year, month):
    """Return (start, end) date inclusive."""
    last = monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last)


# ─────────────────────────────────── GSTR-1 ─────────────────────────────────────

def build_gstr1(*, hospital, year, month):
    """Build a GSTR-1 outward-supplies summary for the given month.

    Excludes CANCELLED invoices. Considers Invoices in PAID/PARTIAL/PENDING/REFUNDED
    states. (For India compliance, sales are recorded at invoice issue date,
    not at payment receipt — this matches the bill_date used here.)

    Returns a dict with sections:
      gstin, period, b2cs, b2cl, hsn_summary, docs, total
    """
    start, end = _month_range(year, month)
    invs = Invoice.objects.filter(
        hospital=hospital,
        bill_date__range=(start, end),
    ).exclude(status__in=["DRAFT", "CANCELLED"])

    items = InvoiceItem.objects.filter(invoice__in=invs).select_related("invoice")

    # B2C summary: state + GST rate aggregation
    # key = (place_of_supply, gst_rate, supply_type) → totals
    b2cs = defaultdict(lambda: {
        "taxable_value": Decimal("0"),
        "cgst": Decimal("0"),
        "sgst": Decimal("0"),
        "igst": Decimal("0"),
    })
    b2cl = []  # list of large inter-state invoices (>2.5L)

    # HSN summary: hsn_code → counts + totals
    hsn = defaultdict(lambda: {
        "qty": Decimal("0"),
        "taxable_value": Decimal("0"),
        "cgst": Decimal("0"),
        "sgst": Decimal("0"),
        "igst": Decimal("0"),
    })

    for it in items:
        inv = it.invoice
        if inv.status == "DRAFT":
            continue
        # Apportion item GST to CGST/SGST/IGST based on invoice gst_split
        rate = Decimal(str(it.gst_rate))
        sub = Decimal(str(it.subtotal))
        gst = Decimal(str(it.gst_amount))
        cgst = sgst = igst = Decimal("0")
        if inv.gst_split == "INTRA":
            cgst = sgst = _q(gst / 2)
        elif inv.gst_split == "INTER":
            igst = _q(gst)

        pos_state = inv.patient_state or inv.hospital_state or "Unknown"

        # B2CL: inter-state invoice value > 2.5L
        if inv.gst_split == "INTER" and Decimal(str(inv.total_amount)) > Decimal("250000"):
            b2cl.append({
                "invoice_no": inv.code,
                "invoice_date": inv.bill_date.isoformat(),
                "invoice_value": str(_q(inv.total_amount)),
                "place_of_supply": pos_state,
                "rate": str(rate),
                "taxable_value": str(_q(sub)),
                "igst": str(igst),
            })
        else:
            key = (pos_state, str(rate), inv.gst_split)
            b2cs[key]["taxable_value"] += sub
            b2cs[key]["cgst"] += cgst
            b2cs[key]["sgst"] += sgst
            b2cs[key]["igst"] += igst

        # HSN
        hsn_code = it.hsn_code or "9993"
        hsn[(hsn_code, str(rate))]["qty"] += Decimal(str(it.quantity))
        hsn[(hsn_code, str(rate))]["taxable_value"] += sub
        hsn[(hsn_code, str(rate))]["cgst"] += cgst
        hsn[(hsn_code, str(rate))]["sgst"] += sgst
        hsn[(hsn_code, str(rate))]["igst"] += igst

    # Format outputs
    b2cs_out = [
        {
            "place_of_supply": pos,
            "rate": rate,
            "supply_type": split,
            "taxable_value": str(_q(v["taxable_value"])),
            "cgst": str(_q(v["cgst"])),
            "sgst": str(_q(v["sgst"])),
            "igst": str(_q(v["igst"])),
        }
        for (pos, rate, split), v in sorted(b2cs.items())
    ]
    hsn_out = [
        {
            "hsn_code": code,
            "rate": rate,
            "quantity": str(_q(v["qty"])),
            "taxable_value": str(_q(v["taxable_value"])),
            "cgst": str(_q(v["cgst"])),
            "sgst": str(_q(v["sgst"])),
            "igst": str(_q(v["igst"])),
            "total_value": str(_q(
                v["taxable_value"] + v["cgst"] + v["sgst"] + v["igst"]
            )),
        }
        for (code, rate), v in sorted(hsn.items())
    ]

    # Docs section
    docs = {
        "invoices_issued": invs.count(),
        "invoices_cancelled": Invoice.objects.filter(
            hospital=hospital, bill_date__range=(start, end), status="CANCELLED",
        ).count(),
    }
    try:
        from apps.billing.models import Refund
        docs["refunds_processed"] = Refund.objects.filter(
            hospital=hospital, status="PROCESSED",
            processed_at__date__range=(start, end),
        ).count()
    except Exception:
        docs["refunds_processed"] = 0

    # Grand totals
    total_taxable = sum((Decimal(r["taxable_value"]) for r in b2cs_out),
                        Decimal("0")) + \
                    sum((Decimal(r["taxable_value"]) for r in b2cl),
                        Decimal("0"))
    total_cgst = sum((Decimal(r["cgst"]) for r in b2cs_out), Decimal("0"))
    total_sgst = sum((Decimal(r["sgst"]) for r in b2cs_out), Decimal("0"))
    total_igst = (sum((Decimal(r["igst"]) for r in b2cs_out), Decimal("0"))
                  + sum((Decimal(r["igst"]) for r in b2cl), Decimal("0")))

    return {
        "gstin": getattr(hospital, "gstin", ""),
        "fp": f"{month:02d}{year}",   # filing period 'MMYYYY' per GSTN spec
        "period": f"{year}-{month:02d}",
        "period_label": start.strftime("%B %Y"),
        "b2cs": b2cs_out,
        "b2cl": b2cl,
        "hsn_summary": hsn_out,
        "docs": docs,
        "totals": {
            "taxable_value": str(_q(total_taxable)),
            "cgst": str(_q(total_cgst)),
            "sgst": str(_q(total_sgst)),
            "igst": str(_q(total_igst)),
            "total_tax": str(_q(total_cgst + total_sgst + total_igst)),
            "grand_total": str(_q(
                total_taxable + total_cgst + total_sgst + total_igst,
            )),
        },
    }


# ─────────────────────────────────── GSTR-3B ────────────────────────────────────

def build_gstr3b(*, hospital, year, month):
    """Build the GSTR-3B summary for the given month.

    Returns a dict in a structure resembling the GSTN JSON schema:
      sec_3_1 = outward supplies
        a) outward taxable supplies (other than zero rated)
        b) zero-rated (export) — N/A for hospitals
        c) other outward supplies (nil-rated, exempted)
        d) inward (reverse charge) — typically not applicable
        e) non-GST outward — typically not applicable
      sec_4 = ITC (we report 0 unless we model purchases — out of scope here)
      sec_5 = exempt/nil/non-gst inward — not modelled
      sec_6_1 = tax payable summary (CGST/SGST/IGST)
    """
    start, end = _month_range(year, month)
    invs = Invoice.objects.filter(
        hospital=hospital,
        bill_date__range=(start, end),
    ).exclude(status__in=["DRAFT", "CANCELLED"])

    items = InvoiceItem.objects.filter(invoice__in=invs).select_related("invoice")

    taxable_value = Decimal("0")
    cgst = sgst = igst = Decimal("0")
    nil_rated = Decimal("0")  # gst_rate = 0%

    for it in items:
        inv = it.invoice
        sub = Decimal(str(it.subtotal))
        rate = Decimal(str(it.gst_rate))
        gst = Decimal(str(it.gst_amount))

        if rate == Decimal("0"):
            # Nil-rated / exempt outward supply (most diagnostic services)
            nil_rated += sub
            continue

        taxable_value += sub
        if inv.gst_split == "INTRA":
            cgst += gst / 2
            sgst += gst / 2
        elif inv.gst_split == "INTER":
            igst += gst

    return {
        "gstin": getattr(hospital, "gstin", ""),
        "ret_period": f"{month:02d}{year}",
        "period": f"{year}-{month:02d}",
        "period_label": start.strftime("%B %Y"),
        "sec_3_1": {
            # (a) Outward taxable supplies
            "outward_taxable": {
                "taxable_value": str(_q(taxable_value)),
                "cgst": str(_q(cgst)),
                "sgst": str(_q(sgst)),
                "igst": str(_q(igst)),
                "cess": "0.00",
            },
            # (b) Zero-rated (exports) — N/A
            "zero_rated": {
                "taxable_value": "0.00",
                "igst": "0.00",
                "cess": "0.00",
            },
            # (c) Other (nil-rated, exempted)
            "other_outward": {
                "taxable_value": str(_q(nil_rated)),
                "cgst": "0.00", "sgst": "0.00", "igst": "0.00", "cess": "0.00",
            },
            # (d) inward reverse charge
            "inward_reverse": {
                "taxable_value": "0.00",
                "cgst": "0.00", "sgst": "0.00", "igst": "0.00", "cess": "0.00",
            },
            # (e) non-GST outward
            "non_gst_outward": {"taxable_value": "0.00"},
        },
        "sec_4_itc": {
            "itc_available": {"cgst": "0.00", "sgst": "0.00", "igst": "0.00"},
            "note": "ITC tracking requires purchase ledger — not in scope of HMS billing module.",
        },
        "sec_6_1_tax_payable": {
            "cgst": str(_q(cgst)),
            "sgst": str(_q(sgst)),
            "igst": str(_q(igst)),
            "total": str(_q(cgst + sgst + igst)),
        },
        "summary": {
            "total_outward_taxable": str(_q(taxable_value)),
            "total_outward_exempt": str(_q(nil_rated)),
            "total_invoices": invs.count(),
            "total_tax_payable": str(_q(cgst + sgst + igst)),
        },
    }


# ───────────────────────────────── Excel export ─────────────────────────────────

def build_gstr_workbook(*, hospital, year, month):
    """Build an .xlsx workbook with both GSTR-1 and GSTR-3B sheets. Returns bytes."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    import io

    g1 = build_gstr1(hospital=hospital, year=year, month=month)
    g3 = build_gstr3b(hospital=hospital, year=year, month=month)

    wb = Workbook()

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="0c4a6e")
    section_font = Font(bold=True, color="0c4a6e", size=12)
    thin = Side(border_style="thin", color="cbd5e1")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def style_header(row):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"GST Returns — {g1['period_label']}"
    ws["A1"].font = Font(bold=True, size=14, color="0c4a6e")
    ws["A2"] = f"GSTIN: {g1['gstin']}"
    ws["A3"] = f"Period: {g1['period_label']}"

    ws["A5"] = "Section"
    ws["B5"] = "Value"
    style_header(ws[5])

    rows = [
        ("Outward Taxable Supplies", g3["sec_3_1"]["outward_taxable"]["taxable_value"]),
        ("Outward Exempt / Nil-rated", g3["sec_3_1"]["other_outward"]["taxable_value"]),
        ("CGST Payable", g3["sec_6_1_tax_payable"]["cgst"]),
        ("SGST Payable", g3["sec_6_1_tax_payable"]["sgst"]),
        ("IGST Payable", g3["sec_6_1_tax_payable"]["igst"]),
        ("Total Tax Payable", g3["sec_6_1_tax_payable"]["total"]),
        ("Invoices Issued", str(g1["docs"]["invoices_issued"])),
        ("Invoices Cancelled", str(g1["docs"]["invoices_cancelled"])),
        ("Refunds Processed", str(g1["docs"]["refunds_processed"])),
    ]
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=f"₹ {v}")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # ── Sheet 2: GSTR-1 B2CS ──
    ws2 = wb.create_sheet("GSTR-1 B2CS")
    ws2["A1"] = "B2C (Small) — State+Rate Summary"
    ws2["A1"].font = section_font
    headers = ["Place of Supply", "Rate (%)", "Supply Type",
               "Taxable Value", "CGST", "SGST", "IGST"]
    for j, h in enumerate(headers, start=1):
        ws2.cell(row=3, column=j, value=h)
    style_header(ws2[3])
    for i, r in enumerate(g1["b2cs"], start=4):
        ws2.cell(row=i, column=1, value=r["place_of_supply"])
        ws2.cell(row=i, column=2, value=r["rate"])
        ws2.cell(row=i, column=3, value=r["supply_type"])
        ws2.cell(row=i, column=4, value=float(r["taxable_value"]))
        ws2.cell(row=i, column=5, value=float(r["cgst"]))
        ws2.cell(row=i, column=6, value=float(r["sgst"]))
        ws2.cell(row=i, column=7, value=float(r["igst"]))
    for col_letter, width in zip("ABCDEFG", [22, 10, 14, 16, 14, 14, 14]):
        ws2.column_dimensions[col_letter].width = width

    # ── Sheet 3: GSTR-1 HSN ──
    ws3 = wb.create_sheet("GSTR-1 HSN")
    ws3["A1"] = "HSN-wise Summary (mandatory)"
    ws3["A1"].font = section_font
    headers = ["HSN Code", "Rate (%)", "Quantity", "Taxable Value",
               "CGST", "SGST", "IGST", "Total Value"]
    for j, h in enumerate(headers, start=1):
        ws3.cell(row=3, column=j, value=h)
    style_header(ws3[3])
    for i, r in enumerate(g1["hsn_summary"], start=4):
        ws3.cell(row=i, column=1, value=r["hsn_code"])
        ws3.cell(row=i, column=2, value=r["rate"])
        ws3.cell(row=i, column=3, value=float(r["quantity"]))
        ws3.cell(row=i, column=4, value=float(r["taxable_value"]))
        ws3.cell(row=i, column=5, value=float(r["cgst"]))
        ws3.cell(row=i, column=6, value=float(r["sgst"]))
        ws3.cell(row=i, column=7, value=float(r["igst"]))
        ws3.cell(row=i, column=8, value=float(r["total_value"]))
    for col_letter, width in zip("ABCDEFGH", [12, 10, 12, 16, 14, 14, 14, 16]):
        ws3.column_dimensions[col_letter].width = width

    # ── Sheet 4: GSTR-3B ──
    ws4 = wb.create_sheet("GSTR-3B")
    ws4["A1"] = "GSTR-3B — Monthly Summary Return"
    ws4["A1"].font = section_font

    ws4["A3"] = "3.1 Outward Supplies"
    ws4["A3"].font = Font(bold=True, color="0c4a6e")
    ws4["A4"] = "(a) Outward Taxable"
    ws4["B4"] = float(g3["sec_3_1"]["outward_taxable"]["taxable_value"])
    ws4["A5"] = "(c) Other Outward (Exempt/Nil)"
    ws4["B5"] = float(g3["sec_3_1"]["other_outward"]["taxable_value"])

    ws4["A7"] = "6.1 Tax Payable"
    ws4["A7"].font = Font(bold=True, color="0c4a6e")
    ws4["A8"] = "CGST"
    ws4["B8"] = float(g3["sec_6_1_tax_payable"]["cgst"])
    ws4["A9"] = "SGST"
    ws4["B9"] = float(g3["sec_6_1_tax_payable"]["sgst"])
    ws4["A10"] = "IGST"
    ws4["B10"] = float(g3["sec_6_1_tax_payable"]["igst"])
    ws4["A11"] = "TOTAL"
    ws4["A11"].font = Font(bold=True)
    ws4["B11"] = float(g3["sec_6_1_tax_payable"]["total"])
    ws4["B11"].font = Font(bold=True)

    ws4.column_dimensions["A"].width = 36
    ws4.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
