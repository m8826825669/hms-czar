"""Import patients (or other entities) from an Excel file.

Usage:
    python manage.py import_excel /path/to/patients.xlsx --entity patients
    python manage.py import_excel /path/to/staff.xlsx --entity staff --dry-run

Expected sheets / columns (configurable - this is the canonical spec for Phase 0):

  patients sheet columns:
    first_name, middle_name, last_name, dob (YYYY-MM-DD), gender (M/F/O),
    blood_group, phone, email, address_line1, city, state, pincode,
    aadhaar_last4, allergies (comma-separated), chronic_conditions

  staff sheet columns:
    username, employee_code, first_name, last_name, email, phone,
    designation, role_code (SUPER_ADMIN / DOCTOR / NURSE / ...)

You can extend ENTITY_HANDLERS to add more entity types.
"""
from __future__ import annotations
from datetime import datetime
from pathlib import Path
import secrets
import string

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.accounts.models import User, Role, UserRole
from apps.core.models import Hospital, Patient


def _generate_password(length: int = 12) -> str:
    alphabet = string.ascii_letters + string.digits + "!@#$%"
    return "".join(secrets.choice(alphabet) for _ in range(length))


def import_patients(rows, hospital, *, dry_run: bool):
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):  # row 1 is header
        try:
            phone = str(row.get("phone", "")).strip()
            if not phone:
                errors.append((i, "Missing phone"))
                continue
            # Skip duplicates by phone
            if Patient.objects.filter(hospital=hospital, phone=phone).exists():
                skipped += 1
                continue

            dob_raw = row.get("dob")
            if isinstance(dob_raw, datetime):
                dob = dob_raw.date()
            else:
                dob = datetime.strptime(str(dob_raw), "%Y-%m-%d").date()

            data = {
                "hospital": hospital,
                "first_name": str(row.get("first_name", "")).strip(),
                "middle_name": str(row.get("middle_name", "") or "").strip(),
                "last_name": str(row.get("last_name", "") or "").strip(),
                "dob": dob,
                "gender": str(row.get("gender", "M")).strip().upper()[:1],
                "blood_group": str(row.get("blood_group", "UNK")).strip().upper() or "UNK",
                "phone": phone,
                "email": str(row.get("email", "") or "").strip(),
                "address_line1": str(row.get("address_line1", "") or "").strip(),
                "city": str(row.get("city", "") or "").strip(),
                "state": str(row.get("state", "") or "").strip(),
                "pincode": str(row.get("pincode", "") or "").strip(),
                "aadhaar_last4": str(row.get("aadhaar_last4", "") or "").strip()[:4],
                "allergies": [
                    {"substance": s.strip(), "severity": "unknown"}
                    for s in str(row.get("allergies", "") or "").split(",") if s.strip()
                ],
                "chronic_conditions": [
                    s.strip() for s in str(row.get("chronic_conditions", "") or "").split(",") if s.strip()
                ],
            }
            if dry_run:
                created += 1
                continue
            patient = Patient(**data)
            patient.mrn = Patient.generate_mrn(hospital)
            patient.save()
            created += 1
        except Exception as e:
            errors.append((i, str(e)))
    return created, skipped, errors


def import_staff(rows, hospital, *, dry_run: bool):
    created, skipped, errors = 0, 0, []
    for i, row in enumerate(rows, start=2):
        try:
            username = str(row.get("username", "")).strip()
            if not username:
                errors.append((i, "Missing username"))
                continue
            if User.objects.filter(username=username).exists():
                skipped += 1
                continue

            password = _generate_password()
            data = {
                "username": username,
                "email": str(row.get("email", "") or "").strip(),
                "first_name": str(row.get("first_name", "")).strip(),
                "last_name": str(row.get("last_name", "") or "").strip(),
                "phone": str(row.get("phone", "") or "").strip(),
                "employee_code": str(row.get("employee_code", "") or "").strip(),
                "designation": str(row.get("designation", "") or "").strip(),
                "hospital": hospital,
                "must_change_password": True,
            }
            role_code = str(row.get("role_code", "") or "").strip().upper()
            if dry_run:
                created += 1
                continue

            user = User(**data)
            user.set_password(password)
            user.save()
            if role_code:
                role = Role.objects.filter(code=role_code).first()
                if role:
                    UserRole.objects.create(user=user, role=role)
            created += 1
            print(f"  [{username}] initial password = {password}")
        except Exception as e:
            errors.append((i, str(e)))
    return created, skipped, errors


ENTITY_HANDLERS = {
    "patients": import_patients,
    "staff": import_staff,
}


class Command(BaseCommand):
    help = "Import data from an Excel (.xlsx) file into HMS."

    def add_arguments(self, parser):
        parser.add_argument("file_path", help="Path to the .xlsx file")
        parser.add_argument("--entity", choices=ENTITY_HANDLERS.keys(), required=True)
        parser.add_argument("--sheet", default=None, help="Sheet name (defaults to first sheet)")
        parser.add_argument("--dry-run", action="store_true",
                            help="Validate without writing to DB")
        parser.add_argument("--hospital-code", default=None,
                            help="Hospital code (defaults to HMS_DEFAULT_HOSPITAL_CODE)")

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError as e:
            raise CommandError("openpyxl not installed") from e

        file_path = Path(opts["file_path"])
        if not file_path.exists():
            raise CommandError(f"File not found: {file_path}")

        hospital_code = opts["hospital_code"] or settings.HMS_DEFAULT_HOSPITAL_CODE
        hospital = Hospital.objects.filter(code=hospital_code).first()
        if not hospital:
            raise CommandError(f"Hospital '{hospital_code}' not found. Run seed_initial first.")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[opts["sheet"]] if opts["sheet"] else wb[wb.sheetnames[0]]

        # First non-empty row is the header
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower() if h else "" for h in next(rows_iter)]
        rows = [dict(zip(headers, row)) for row in rows_iter
                if any(cell is not None for cell in row)]

        self.stdout.write(f"File: {file_path}, sheet: {ws.title}, rows: {len(rows)}")
        self.stdout.write(f"Headers: {headers}")
        self.stdout.write(f"Mode: {'DRY RUN' if opts['dry_run'] else 'WRITE'}")

        handler = ENTITY_HANDLERS[opts["entity"]]
        with transaction.atomic():
            created, skipped, errors = handler(rows, hospital, dry_run=opts["dry_run"])
            if opts["dry_run"]:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS(f"Created: {created}"))
        self.stdout.write(self.style.WARNING(f"Skipped (duplicates): {skipped}"))
        if errors:
            self.stdout.write(self.style.ERROR(f"Errors: {len(errors)}"))
            for r, msg in errors[:20]:
                self.stdout.write(f"  Row {r}: {msg}")
            if len(errors) > 20:
                self.stdout.write(f"  ... and {len(errors) - 20} more")
